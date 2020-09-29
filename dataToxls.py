# !/usr/bin/python
# coding:utf-8

import openpyxl
from openpyxl import load_workbook
from openpyxl.workbook import Workbook, defined_name
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import colors, Font, Fill, NamedStyle
from openpyxl.styles import PatternFill, Border, Side, Alignment
import json
import mysql.connector
import io

def takeThird(elem):
    return elem[2]

def convert(cnt):
    cnt -= 65
    if cnt > 25:
        cntC = chr(64 + cnt // 26) + chr(65 + (cnt % 26))
    else:
        cntC = chr(65 + (cnt % 26))
    return cntC

arkdb = mysql.connector.connect(
    host="192.168.168.146",
    user="arkdeve",
    password="BGuy6013@",
    database="arknightDB",
    buffered=True
)
cursor = arkdb.cursor()
cursor.execute("SELECT char_id, job, star, charName FROM charSummary")
datas = cursor.fetchall()
cursor.close()

cursor = arkdb.cursor()
cursor.execute("SELECT char_id FROM charSummary ORDER BY char_id DESC")
lastID = cursor.fetchone()
cursor.close()
lastID = lastID[0]
# print(lastID[0])

left, right, top, bottom = [Side(style='thin', color='000000')]*4

Job = ["先鋒", "狙擊", "醫療", "術師", "近衛", "重裝", "輔助", "特種"]
category1 = ["pioneer", "sniper", "medic", "caster", "warrior", "tank", "support", "special"]
category2 = ["0_2", "0_3", "0_4", "0_5", "0_6", "0_7", "1_8", "1_9", "1_10", "2_8", "2_9", "2_10", "3_8", "3_9", "3_10"]
category3 = [1, 3, 2, 3, 2, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3]
category4 = ["m1", "m2", "m3", "m4"]

sorted = []
for i in range(0, 8):
    temp = []
    for data in datas:
        if data[1] == category1[i]:
            temp.append(data)
    temp.sort(key=takeThird, reverse=True)
    sorted.append(temp)
# print(sorted)

f = io.open("check.json", "r", encoding="utf-8")
checks = json.load(f)
f.close()

wb = Workbook()

wc = wb['Sheet']
wc.title = 'Char'
ws = wb.create_sheet('Summary')
wd = wb.create_sheet('Skill')
we = wb.create_sheet('Elite')
wm = wb.create_sheet('Statistics')

# 角色簡介
for i in range(1, 9):
    ws['A' + str(i)].value = Job[i-1]
    ws['A' + str(i)].alignment = Alignment(horizontal='center', vertical='center')

for j in range(0, 8):
    cnt = 2
    star = 6
    temp = 0
    ws[convert(66 + j*3 + 2) + '1'].value = "----6----"
    ws[convert(66 + j*3 + 2) + '1'].alignment = Alignment(horizontal='center', vertical='center')
    for char in sorted[j]:
        for i in range(0, 3):
            if char[2] != star:
                star = char[2]
                ws[convert(66 + j*3 + 2) + str(cnt + temp)].value = "----" + str(star) + "----"
                ws[convert(66 + j*3 + 2) + str(cnt + temp)].alignment = Alignment(horizontal='center', vertical='center')
                temp += 1
            if i > 0:
                ws[convert(66 + j*3 + i) + str(cnt + temp)].value = char[i+1]
                ws[convert(66 + j*3 + i) + str(cnt + temp)].alignment = Alignment(horizontal='center', vertical='center')
            else:
                ws[convert(66 + j*3 + i) + str(cnt + temp)].value = char[i]
                ws[convert(66 + j*3 + i) + str(cnt + temp)].alignment = Alignment(horizontal='center', vertical='center')
            if i%3 == 0:
                ws.column_dimensions[convert(66 + j*3 + i)].width = 5
            elif i%3 == 1:
                ws.column_dimensions[convert(66 + j*3 + i)].width = 7
            else:
                ws.column_dimensions[convert(66 + j*3 + i)].width = 11
        cnt += 1

# 技能需求
temp = 66
for i in range(0, 15):
    for j in range(0, category3[i]):
        wd[convert(temp + j*2) + '1'].value = 'S' + category2[i] + '_' + category4[j] + '_n'
        wd[convert(temp + j*2) + '1'].alignment = Alignment(horizontal='center', vertical='center')
        wd[convert(temp + j*2 + 1) + '1'].value = 'S' + category2[i] + '_' + category4[j] + '_q'
        wd[convert(temp + j*2 + 1) + '1'].alignment = Alignment(horizontal='center', vertical='center')
    temp += category3[i]*2

for i in range(1, lastID):
    check = checks[str(i)]
    wd['A' + str(i+1)].value = check['name']
    wd['A' + str(i+1)].alignment = Alignment(horizontal='center', vertical='center')
    sk = check['demand']['skill']
    temp = 66
    for j in range(0, 15):
        for k in range(0, category3[j]):
            try:
                wd[convert(temp + k*2) + str(i+1)].value = sk[category2[j]]['materials'][category4[k]]['name']
                wd[convert(temp + k*2) + str(i+1)].alignment = Alignment(horizontal='center', vertical='center')
                wd.column_dimensions[convert(temp + k*2)].width = 13
                wd[convert(temp + k*2 + 1) + str(i+1)].value = sk[category2[j]]['materials'][category4[k]]['quantity']
                wd[convert(temp + k*2 + 1) + str(i+1)].alignment = Alignment(horizontal='center', vertical='center')
                wd.column_dimensions[convert(temp + k*2 + 1)].width = 11
            except:
                break
        temp += category3[j]*2
wd.column_dimensions['A'].width = 11

# 菁英化需求
for i in range(0, 2):
    for j in range(0, 4):
        we[convert((66 + i*8) + j*2) + '1'].value = 'E' + str(i+1) + '_' + category4[j] + '_n'
        we[convert((66 + i*8) + j*2) + '1'].alignment = Alignment(horizontal='center', vertical='center')
        we[convert((66 + i*8) + j*2 + 1) + '1'].value = 'E' + str(i+1) + '_' + category4[j] + '_q'
        we[convert((66 + i*8) + j*2 + 1) + '1'].alignment = Alignment(horizontal='center', vertical='center')

for i in range(1, lastID):
    check = checks[str(i)]
    we['A' + str(i+1)].value = check['name']
    we['A' + str(i+1)].alignment = Alignment(horizontal='center', vertical='center')
    el = check['demand']['elite']
    for j in range(0, 2):
        for k in range(0, 4):
            try:
                we[convert(66 + j*8 + k*2) + str(i+1)].value = el[str(j+1)]['materials'][category4[k]]['name']
                we[convert(66 + j*8 + k*2) + str(i+1)].alignment = Alignment(horizontal='center', vertical='center')
                we.column_dimensions[convert(66 + j*8 + k*2)].width = 13
                we[convert((66 + j*8 + k*2) + 1) + str(i+1)].value = el[str(j+1)]['materials'][category4[k]]['quantity']
                we[convert((66 + j*8 + k*2) + 1) + str(i+1)].alignment = Alignment(horizontal='center', vertical='center')
                we.column_dimensions[convert((66 + j*8 + k*2) + 1)].width = 11
            except:
                break
we.column_dimensions['A'].width = 11

# 下拉式選單
new_range = defined_name.DefinedName('職業', attr_text='Summary!$A$1:$A$8')
wb.defined_names.append(new_range)

for i in range(0, 8):
    for j in range(2, 100):
        if ws[convert(68 + i*3) + str(j)].value is None:
            break
    j -= 1

    new_range = defined_name.DefinedName(Job[i], attr_text='Summary!$' + convert(68 + i*3) + '$1:$' + convert(68 + i*3) + '$' + str(j))
    wb.defined_names.append(new_range)

wc['B2'].value = '職業'
wc['B2'].alignment = Alignment(horizontal='center', vertical='center')
wc['B2'].border = Border(left=left, right=right, top=top, bottom=bottom)

data_val = DataValidation(type='list', formula1='=職業')
wc.add_data_validation(data_val)
data_val.add(wc['C2'])
wc['C2'].alignment = Alignment(horizontal='center', vertical='center')
wc['C2'].border = Border(left=left, right=right, top=top, bottom=bottom)

wc['B3'].value = '角色'
wc['B3'].alignment = Alignment(horizontal='center', vertical='center')
wc['B3'].border = Border(left=left, right=right, top=top, bottom=bottom)

data_val = DataValidation(type='list', formula1='=INDIRECT(C2)')
wc.add_data_validation(data_val)
data_val.add(wc['C3'])
wc['C3'].alignment = Alignment(horizontal='center', vertical='center')
wc['C3'].border = Border(left=left, right=right, top=top, bottom=bottom)

wc.column_dimensions['D'].width = 2

# 資料顯示 - 技能
init = 2
for i in range(0, 6):
    wc.merge_cells('E' + str(init) + ':E' + str(init + category3[i] - 1))
    wc['E' + str(init)].value = '技能等級' + str(i+1)
    wc['E' + str(init)].alignment = Alignment(horizontal='center', vertical='center')
    init += category3[i]

for i in range(2, init):
    wc['F' + str(i)].value = '=IF(IFERROR(VLOOKUP(C3, Skill!$A$2:$CE$160, ' + str(i*2 - 2) + ', FALSE), "")=0, "", IFERROR(VLOOKUP(C3, Skill!$A$2:$CE$160, ' + str(i*2 - 2) + ', FALSE), ""))'
    wc['G' + str(i)].value = '=IF(IFERROR(VLOOKUP(C3, Skill!$A$2:$CE$160, ' + str(i*2 - 1) + ', FALSE), "")=0, "", IFERROR(VLOOKUP(C3, Skill!$A$2:$CE$160, ' + str(i*2 - 1) + ', FALSE), ""))'

for j in range(0, 3):
    for i in range(2, init):
        wc[convert(69 + j) + str(i)].border = Border(left=left, right=right, top=top, bottom=bottom)

wc.column_dimensions['E'].width = 11
wc.column_dimensions['F'].width = 13
wc.column_dimensions['G'].width = 7

wc.column_dimensions['H'].width = 1

temp = 2
for i in range(1, 4):
    wc.merge_cells('I' + str(temp) + ':I' + str(temp + 8))
    wc['I' + str(temp)].value = '技能' + str(i)
    wc['I' + str(temp)].alignment = Alignment(horizontal='center', vertical='center')
    init = 1
    for j in range(1, 4):
        wc.merge_cells('J' + str(temp) + ':J' + str(temp + 2))
        wc['J' + str(temp)].value = '專' + str(init)
        wc['J' + str(temp)].alignment = Alignment(horizontal='center', vertical='center')
        temp += 3
        init += 1

for i in range(2, temp):
    wc['K' + str(i)].value = '=IF(IFERROR(VLOOKUP(C3, Skill!$A$2:$CE$160, ' + str(i*2 + 26) + ', FALSE), "")=0, "", IFERROR(VLOOKUP(C3, Skill!$A$2:$CE$160, ' + str(i*2 + 26) + ', FALSE), ""))'
    wc['L' + str(i)].value = '=IF(IFERROR(VLOOKUP(C3, Skill!$A$2:$CE$160, ' + str(i*2 + 27) + ', FALSE), "")=0, "", IFERROR(VLOOKUP(C3, Skill!$A$2:$CE$160, ' + str(i*2 + 27) + ', FALSE), ""))'

for j in range(0, 4):
    for i in range(2, temp):
        wc[convert(73 + j) + str(i)].border = Border(left=left, right=right, top=top, bottom=bottom)

wc.column_dimensions['I'].width = 7
wc.column_dimensions['J'].width = 7
wc.column_dimensions['K'].width = 13
wc.column_dimensions['L'].width = 11

wc.column_dimensions['M'].width = 2

# 資料顯示 - 菁英化
init = 2
for i in range(0, 2):
    wc.merge_cells('N' + str(init) + ':N' + str(init + 3))
    wc['N' + str(init)].value = '菁英' + str(i+1)
    wc['N' + str(init)].alignment = Alignment(horizontal='center', vertical='center')
    init += 4

for i in range(2, init):
    wc['O' + str(i)].value = '=IF(IFERROR(VLOOKUP(C3, Elite!$A$2:$Q$160, ' + str(i*2 - 2) + ', FALSE), "")=0, "", IFERROR(VLOOKUP(C3, Elite!$A$2:$Q$160, ' + str(i*2 - 2) + ', FALSE), ""))'
    wc['P' + str(i)].value = '=IF(IFERROR(VLOOKUP(C3, Elite!$A$2:$Q$160, ' + str(i*2 - 1) + ', FALSE), "")=0, "", IFERROR(VLOOKUP(C3, Elite!$A$2:$Q$160, ' + str(i*2 - 1) + ', FALSE), ""))'
  
for j in range(0, 3):
    for i in range(2, init):
        wc[convert(78 + j) + str(i)].border = Border(left=left, right=right, top=top, bottom=bottom)

wc.column_dimensions['N'].width = 7
wc.column_dimensions['O'].width = 13
wc.column_dimensions['P'].width = 7

# 資料儲存 - 物資總計
cursor = arkdb.cursor()
cursor.execute("SELECT m_name FROM material")
materials = cursor.fetchall()
cursor.close()

wm['B1'].value = '庫存'
wm['B1'].alignment = Alignment(horizontal='center', vertical='center')
wm['B1'].border = Border(left=left, right=right, top=top, bottom=bottom)
wm['C1'].value = '需求'
wm['C1'].alignment = Alignment(horizontal='center', vertical='center')
wm['C1'].border = Border(left=left, right=right, top=top, bottom=bottom)

init = 1
for material in materials:
    wm['A' + str(init+1)].value = material[0]
    wm['A' + str(init+1)].alignment = Alignment(horizontal='center', vertical='center')
    for i in range(0, 3):
        wm[convert(65+i) + str(init+1)].border = Border(left=left, right=right, top=top, bottom=bottom)
    init += 1

wm.column_dimensions['A'].width = 15

wb.save('Test.xlsx')